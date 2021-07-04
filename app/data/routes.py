from datetime import datetime
from flask import render_template, flash, redirect, url_for, request, g, \
	jsonify, current_app, Response, stream_with_context, make_response, session
from flask_login import current_user, login_required
from app import db, csrf
from app.models import User, Post, Program, Message, Notification, Interview, Interview_Date, Test, Specialty, Chat, Interview_Impression, Thread, PostInterviewCommunication
import logging
import re
import pandas as pd
import datetime as dt
from datetime import datetime
import dateutil.parser
import os
import json
from app.data import bp
import urllib.request
import app
import openpyxl as openpyxl

@bp.route('/delete_specialties')
def delete_specialties():
	if current_user.admin:
		Specialty.query.delete()
		db.session.commit()
	return redirect(url_for('main.index'))

@bp.route('/create_specialties')
def create_specialties():
	if current_user.admin:
		with current_app.open_resource('static/data/specialty.json') as f:
			data = json.load(f)
			for i in data:
				db.session.add(Specialty(name=i['specialty'][0]))
				db.session.commit()
		return str(data)
	return redirect(url_for('main.index'))

@bp.route('/delete_programs')
def delete_programs():
	if current_user.admin:
		Program.query.delete()
		db.session.commit()
	return redirect(url_for('main.index'))

@bp.route('/upload_surgery', methods=['GET', 'POST'])
@csrf.exempt
def upload_surgery():
	if request.method == 'POST':
		if current_user.admin:
			def generate():
				spec = Specialty.query.filter_by(name='Surgery').first_or_404()
				print(spec.name)
				wb = openpyxl.load_workbook(request.files['file'])
				ws = wb.active
				for row in ws.iter_rows(min_row=1):
					program_name = row[0].value
					program = Program.query.filter_by(name=program_name, specialty_id=spec.id).first_or_404()
					program_offer_dates = row[1].value
					program_interview_dates = row[2].value
					offer_dates = None
					interview_dates = None
					if program_offer_dates:
						offer_dates = list(map(lambda x: Interview(date=datetime.strptime(x, "%m/%d/%Y"), interviewer=program,interviewee=current_user), program_offer_dates.split(",")))
						for offer_date in offer_dates:
							db.session.add(offer_date)
					if program_interview_dates:
						interview_dates = list(map(lambda x: Interview_Date(date=datetime.strptime(x, "%m/%d/%Y"), interviewer=program,interviewee=current_user,full=True), program_interview_dates.split(",")))
						for interview_date in interview_dates:
							db.session.add(interview_date)
					db.session.commit()
					yield(program_name)
			return Response(stream_with_context(generate()))
		return redirect(url_for('main.index'))
	return render_template('upload.html')

@bp.route('/upload_surgery_interview_impressions', methods=['GET', 'POST'])
@csrf.exempt
def upload_surgery_interview_impressions():
	if request.method == 'POST':
		if current_user.admin:
			def generate():
				spec = Specialty.query.filter_by(name='Surgery').first()
				wb = openpyxl.load_workbook(request.files['file'])
				ws = wb.active
				for row in ws.iter_rows(min_row=1):
					program_name = row[0].value
					program = Program.query.filter_by(name=program_name, specialty_id=spec.id).first()
					if program:
						program_interview_impressions = row[1].value
						interview_impressions = None
						if program_interview_impressions:
							interview_impressions = list(map(lambda x: Interview_Impression(body=x, author=current_user, program=program), program_interview_impressions.split("XXXXX")))
							for interview_impression in interview_impressions:
								db.session.add(interview_impression)
						db.session.commit()
						yield(program_name)
			return Response(stream_with_context(generate()))
		return redirect(url_for('main.index'))
	return render_template('upload.html')

@bp.route('/upload_surgery_name_and_shame', methods=['GET', 'POST'])
@csrf.exempt
def upload_surgery_name_and_shame():
	if request.method == 'POST':
		if current_user.admin:
			def generate():
				spec = Specialty.query.filter_by(name='Surgery').first()
				wb = openpyxl.load_workbook(request.files['file'])
				ws = wb.active
				for row in ws.iter_rows(min_row=1):
					program_name = row[0].value
					program = Program.query.filter_by(name=program_name, specialty_id=spec.id).first()
					if program:
						program_interview_impressions = row[1].value
						interview_impressions = None
						if program_interview_impressions:
							interview_impressions = list(map(lambda x: Interview_Impression(body=x, author=current_user, program=program, name_and_shame=True), program_interview_impressions.split("XXXXX")))
							for interview_impression in interview_impressions:
								db.session.add(interview_impression)
						db.session.commit()
						yield(program_name)
			return Response(stream_with_context(generate()))
		return redirect(url_for('main.index'))
	return render_template('upload.html')

@bp.route('/upload_surgery_postiv_communications', methods=['GET', 'POST'])
@csrf.exempt
def upload_surgery_postiv_communications():
	if request.method == 'POST':
		if current_user.admin:
			def generate():
				spec = Specialty.query.filter_by(name='Surgery').first()
				wb = openpyxl.load_workbook(request.files['file'])
				ws = wb.active
				for row in ws.iter_rows(min_row=1):
					program_name = row[0].value
					program = Program.query.filter_by(name=program_name, specialty_id=spec.id).first()
					if program:
						tally = row[1].value
						date = row[2].value
						type_of_communication = row[3].value
						if row[4].value == "Personalized":
							personalized = True
						else:
							personalized = False
						content = row[5].value
						postiv = PostInterviewCommunication(date_of_communication=date, author=current_user, program=program, type_of_communication=type_of_communication, personalized=personalized, content=content)
						db.session.add(postiv)
						db.session.commit()
						yield(program_name)
			return Response(stream_with_context(generate()))
		return redirect(url_for('main.index'))
	return render_template('upload.html')

@bp.route('/upload_surgery_chat', methods=['GET', 'POST'])
@csrf.exempt
def upload_surgery_chat():
	if request.method == 'POST':
		if current_user.admin:
			def generate():
				spec = Specialty.query.filter_by(name='Surgery').first()
				wb = openpyxl.load_workbook(request.files['file'])
				ws = wb.active
				for row in ws.iter_rows(min_row=1):
					if row[0].value:
						thread_body = row[0].value
						thread = Thread(body=thread_body, author=current_user,specialty=spec)
						db.session.add(thread)
						db.session.commit()
						response_list = row[1].value
						if response_list:
							responses = list(map(lambda x: Post(body=x, author=current_user, thread_id=thread.id), response_list.split("XXXXX")))
							for response in responses:
								db.session.add(response)
						db.session.commit()
						yield(str(row[0].row))
			return Response(stream_with_context(generate()))
		return redirect(url_for('main.index'))
	return render_template('upload.html')

@bp.route('/create_programs')
def create_programs():
	def generate(): 
		if current_user.admin:
			with current_app.open_resource('static/data/programs.json') as f:
				data = json.load(f)
				count = 0
				for i in data:
					specialty = Specialty.query.filter_by(name=i['specialty']).first_or_404()
					db.session.add(Program(specialty=specialty, name=next(iter(i['program']), None), city=next(iter(i['city']), None), state=next(iter(i['state']), None), accreditation_id=next(iter(i['accreditation_id']), None),status=next(iter(i['status']), None),url=next(iter(i['url']), None)))
					db.session.commit()
					count = count + 1
					yield(str(count) + " ")
	return Response(stream_with_context(generate()))

@bp.route('/delete_interviews')
def delete_interviews():
	if current_user.admin:
		Interview.query.delete()
		db.session.commit()
	return redirect(url_for('main.index'))

@bp.route('/delete_interview_dates')
def delete_interview_dates():
	if current_user.admin:
		Interview_Date.query.delete()
		db.session.commit()
	return redirect(url_for('main.index'))

@bp.route('/delete_chats')
def delete_chats():
	if current_user.admin:
		Chat.query.delete()
		db.session.commit()
	return redirect(url_for('main.index'))

@bp.route('/delete_interview_impressions')
def delete_interview_impressions():
	if current_user.admin:
		Interview_Impression.query.delete()
		db.session.commit()
	return redirect(url_for('main.index'))

@bp.route('/remove_user_specialties')
def remove_user_specialties():
	if current_user.admin:
		users = User.query.all()
		for user in users:
			user.specialty_id = None
			user.programs = []
		db.session.commit()
		specialties = Specialty.query.all()
		for specialty in specialties:
			specialty.users = []
			specialty.programs = []
		db.session.commit()
		for program in Program.query.all():
			program.users = []
		db.session.commit()
	return redirect(url_for('main.index'))

@bp.route('/add_salary')
def add_salary():
	def generate(): 
		if current_user.admin:
			with current_app.open_resource('static/data/salary.json') as f:
				data = json.load(f)
				count = 0
				for i in data:
					specialty = Specialty.query.filter_by(name=i['specialty']).first_or_404()
					program = Program.query.filter_by(name=i['program'], specialty=specialty).first_or_404()
					found = i['found']
					if found:
						program.salary = ", ".join(i['found'])
						program.salary_url = i['url']
						db.session.commit()
					count = count + 1
					yield(str(count) + " ")
	return Response(stream_with_context(generate()))

@bp.route('/add_images')
def add_images():
	def generate(): 
		if current_user.admin:
			with current_app.open_resource('static/data/images.json') as f:
				data = json.load(f)
				count = 0
				for i in data:
					specialty = Specialty.query.filter_by(name=i['specialty']).first_or_404()
					program = Program.query.filter_by(name=i['program'], specialty=specialty).first_or_404()
					program.imageurl = i['image']
					db.session.commit()
					count = count + 1
					yield(str(count) + " ")
	return Response(stream_with_context(generate()))